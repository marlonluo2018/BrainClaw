"""
Split monthly L&K Learning Report by Sub Team (focal) for Industry & Delivery team.

Usage: py -3 assistant_brain/scripts/split_lk_report_by_focal.py <source_folder>
Example: py -3 assistant_brain/scripts/split_lk_report_by_focal.py "C:/Users/MengNingLuo/Downloads/June"

Source folder should contain the 4 monthly report Excel files.
Output goes to <source_folder>/output/<Month YYYY - SubTeam>.xlsx
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
import re
from datetime import datetime

TEAM_FILTER = "Industry & Delivery"

# File patterns and their raw data sheet configs
# Each pattern tries to match one of the 4 report types by keywords in filename
FILE_PATTERNS = [
    {
        "keyword": "All badge",
        "sheet": "Raw Data",
        "team_col": 5,
        "sub_team_col": 6,
        "output_sheet_name": "All Badge",
    },
    {
        "keyword": "Industry badge",
        "sheet": "YL Report",
        "team_col": 5,
        "sub_team_col": 6,
        "output_sheet_name": "Industry Badge",
    },
    {
        "keyword": "Language",
        "sheet": "final",
        "team_col": 6,
        "sub_team_col": 7,
        "output_sheet_name": "Language",
    },
    {
        "keyword": "T40",
        "sheet": "HC",
        "team_col": 10,
        "sub_team_col": 11,
        "output_sheet_name": "T40",
    },
]


def detect_month_from_folder(source_dir):
    """Try to detect the month name from folder name or file names."""
    folder_name = os.path.basename(source_dir)
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    for m in months:
        if m.lower() in folder_name.lower():
            return m
    # Try from filenames
    for f in os.listdir(source_dir):
        if f.endswith(".xlsx") and not f.startswith("~$"):
            match = re.search(r"(\d{4})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*", f, re.IGNORECASE)
            if match:
                year = match.group(1)
                month_str = match.group(2)
                for m in months:
                    if m.lower().startswith(month_str.lower()):
                        return f"{m} {year}"
    return folder_name


def match_file_to_config(filename, patterns):
    """Match a filename to one of the config patterns."""
    for pattern in patterns:
        if pattern["keyword"].lower() in filename.lower():
            return pattern
    return None


def load_filtered_data(filepath, config):
    """Load data filtered by Team = 'Industry & Delivery', grouped by Sub Team."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if config["sheet"] not in wb.sheetnames:
        print(f"  WARNING: Sheet '{config['sheet']}' not found in {os.path.basename(filepath)}, skipping")
        wb.close()
        return None, {}

    ws = wb[config["sheet"]]
    headers = None
    grouped = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        team = row[config["team_col"]] if config["team_col"] < len(row) else None
        sub_team = row[config["sub_team_col"]] if config["sub_team_col"] < len(row) else None
        if str(team).strip() == TEAM_FILTER:
            st_key = str(sub_team).strip() if sub_team else "Unknown"
            if st_key not in grouped:
                grouped[st_key] = []
            grouped[st_key].append(list(row))

    wb.close()
    return headers, grouped


def main():
    if len(sys.argv) < 2:
        print("Usage: py -3 split_lk_report_by_focal.py <source_folder>")
        print('Example: py -3 split_lk_report_by_focal.py "C:/Users/MengNingLuo/Downloads/June"')
        sys.exit(1)

    source_dir = sys.argv[1]
    if not os.path.isdir(source_dir):
        print(f"ERROR: Folder not found: {source_dir}")
        sys.exit(1)

    output_dir = os.path.join(source_dir, "output")
    os.makedirs(output_dir, exist_ok=True)

    month_label = detect_month_from_folder(source_dir)
    print(f"Report period: {month_label}")
    print(f"Source: {source_dir}")
    print(f"Output: {output_dir}\n")

    # Find and match Excel files
    excel_files = [f for f in os.listdir(source_dir) if f.endswith(".xlsx") and not f.startswith("~$")]
    all_data = {}
    matched = 0

    for filename in excel_files:
        config = match_file_to_config(filename, FILE_PATTERNS)
        if config is None:
            print(f"Skipping (no pattern match): {filename}")
            continue

        filepath = os.path.join(source_dir, filename)
        print(f"Processing: {filename} -> sheet '{config['sheet']}'")
        headers, grouped = load_filtered_data(filepath, config)

        if headers is None:
            continue

        matched += 1
        for sub_team, rows in grouped.items():
            if sub_team not in all_data:
                all_data[sub_team] = {}
            all_data[sub_team][config["output_sheet_name"]] = (headers, rows)
            print(f"  {sub_team}: {len(rows)} rows")

    if matched == 0:
        print("\nERROR: No matching report files found in source folder.")
        sys.exit(1)

    # Write output files
    print(f"\n--- Writing output ---")
    for sub_team, sheets in sorted(all_data.items()):
        out_filename = f"{month_label} - {sub_team}.xlsx"
        out_path = os.path.join(output_dir, out_filename)

        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)

        for sheet_name, (headers, rows) in sheets.items():
            ws = wb_out.create_sheet(title=sheet_name)
            ws.append(headers)
            for row in rows:
                ws.append(row)
            for col_idx, header in enumerate(headers, 1):
                if header:
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(30, len(str(header)) + 2))

        wb_out.save(out_path)
        total_rows = sum(len(r) for _, (_, r) in sheets.items())
        print(f"  {out_filename} ({len(sheets)} sheets, {total_rows} rows)")

    print(f"\nDone! {len(all_data)} focal files created in {output_dir}")


if __name__ == "__main__":
    main()
